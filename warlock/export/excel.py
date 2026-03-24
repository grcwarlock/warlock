"""Excel export for GRC data — findings, control results, POA&Ms, risk register.

Uses openpyxl with professional formatting: header styling, auto-width columns,
conditional formatting, freeze panes, and auto-filters.
"""

from __future__ import annotations

import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, NamedStyle, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from warlock.utils import ensure_aware

log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Shared styles
# ---------------------------------------------------------------------------

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

_GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_GREEN_FONT = Font(color="006100")
_RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_RED_FONT = Font(color="9C0006")
_YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_YELLOW_FONT = Font(color="9C6500")

_DATE_FORMAT = "YYYY-MM-DD HH:MM"
_CURRENCY_FORMAT = "#,##0.00"


def _create_header_style() -> NamedStyle:
    """Create a reusable header named style."""
    style = NamedStyle(name="warlock_header")
    style.font = _HEADER_FONT
    style.fill = _HEADER_FILL
    style.alignment = _HEADER_ALIGNMENT
    return style


def _apply_headers(ws: Worksheet, headers: list[str]) -> None:
    """Write header row with styling and freeze the top row."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _auto_width(ws: Worksheet) -> None:
    """Auto-size columns based on content width, with a reasonable max."""
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                cell_len = len(str(cell.value or ""))
                max_length = max(max_length, cell_len)
            except Exception:
                pass
        # Cap at 50 characters, minimum 10
        adjusted = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[col_letter].width = adjusted


def _add_status_formatting(ws: Worksheet, status_col: int, max_row: int) -> None:
    """Add conditional formatting for status columns."""
    col_letter = get_column_letter(status_col)
    cell_range = f"{col_letter}2:{col_letter}{max_row}"

    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator="equal",
            formula=['"compliant"'],
            fill=_GREEN_FILL,
            font=_GREEN_FONT,
        ),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator="equal",
            formula=['"non_compliant"'],
            fill=_RED_FILL,
            font=_RED_FONT,
        ),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator="equal",
            formula=['"partial"'],
            fill=_YELLOW_FILL,
            font=_YELLOW_FONT,
        ),
    )


def _add_severity_formatting(ws: Worksheet, sev_col: int, max_row: int) -> None:
    """Add conditional formatting for severity columns."""
    col_letter = get_column_letter(sev_col)
    cell_range = f"{col_letter}2:{col_letter}{max_row}"

    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator="equal",
            formula=['"critical"'],
            fill=_RED_FILL,
            font=_RED_FONT,
        ),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator="equal",
            formula=['"high"'],
            fill=PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),
            font=Font(color="CC0000"),
        ),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator="equal",
            formula=['"medium"'],
            fill=_YELLOW_FILL,
            font=_YELLOW_FONT,
        ),
    )


def _iso_str(dt: datetime | None) -> str | None:
    """Convert a datetime to ISO string, handling naive datetimes."""
    if dt is None:
        return None
    dt = ensure_aware(dt)
    return dt.strftime("%Y-%m-%d %H:%M:%S")


# ---------------------------------------------------------------------------
# ExcelExporter
# ---------------------------------------------------------------------------


class ExcelExporter:
    """Export GRC data to formatted Excel workbooks.

    Produces professional-quality .xlsx files with:
    - Bold header row with blue background
    - Column auto-width
    - Conditional formatting for status and severity
    - Frozen header row
    - Auto-filters on all columns
    """

    def export_findings(
        self,
        session: Session,
        filepath: str | Path,
        framework: str | None = None,
        severity: str | None = None,
    ) -> Path:
        """Export findings to an Excel workbook.

        Args:
            session: SQLAlchemy session.
            filepath: Output .xlsx path.
            framework: Optional framework filter.
            severity: Optional severity filter.

        Returns:
            Path to the written file.
        """
        from warlock.db.models import Finding

        query = session.query(Finding)
        if severity:
            query = query.filter(Finding.severity == severity)
        if framework:
            # Filter findings that have control results for this framework
            from warlock.db.models import ControlResult

            finding_ids = (
                session.query(ControlResult.finding_id)
                .filter(ControlResult.framework == framework)
                .distinct()
            )
            query = query.filter(Finding.id.in_(finding_ids))

        query = query.order_by(Finding.ingested_at.desc())
        findings = query.limit(50000).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Findings"

        headers = [
            "ID",
            "Title",
            "Observation Type",
            "Severity",
            "Source",
            "Provider",
            "Resource ID",
            "Resource Type",
            "Account ID",
            "Region",
            "Observed At",
            "Ingested At",
            "SHA256",
        ]
        _apply_headers(ws, headers)

        for row_idx, f in enumerate(findings, 2):
            ws.cell(row=row_idx, column=1, value=f.id)
            ws.cell(row=row_idx, column=2, value=f.title)
            ws.cell(row=row_idx, column=3, value=f.observation_type)
            ws.cell(row=row_idx, column=4, value=f.severity)
            ws.cell(row=row_idx, column=5, value=f.source)
            ws.cell(row=row_idx, column=6, value=f.provider)
            ws.cell(row=row_idx, column=7, value=f.resource_id or "")
            ws.cell(row=row_idx, column=8, value=f.resource_type or "")
            ws.cell(row=row_idx, column=9, value=f.account_id or "")
            ws.cell(row=row_idx, column=10, value=f.region or "")
            ws.cell(row=row_idx, column=11, value=_iso_str(f.observed_at))
            ws.cell(row=row_idx, column=12, value=_iso_str(f.ingested_at))
            ws.cell(row=row_idx, column=13, value=f.sha256)

        max_row = len(findings) + 1
        if max_row > 1:
            _add_severity_formatting(ws, 4, max_row)

        _auto_width(ws)
        ws.auto_filter.ref = f"A1:M{max_row}"

        out = Path(filepath)
        wb.save(str(out))
        log.info("Exported %d findings to %s", len(findings), out)
        return out

    def export_control_results(
        self,
        session: Session,
        filepath: str | Path,
        framework: str | None = None,
    ) -> Path:
        """Export control results with conditional status formatting.

        Args:
            session: SQLAlchemy session.
            filepath: Output .xlsx path.
            framework: Optional framework filter.

        Returns:
            Path to the written file.
        """
        from warlock.db.models import ControlResult

        query = session.query(ControlResult)
        if framework:
            query = query.filter(ControlResult.framework == framework)
        query = query.order_by(ControlResult.framework, ControlResult.control_id)
        results = query.limit(100000).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Control Results"

        headers = [
            "ID",
            "Framework",
            "Control ID",
            "Status",
            "Severity",
            "Assertion",
            "Assertion Passed",
            "AI Confidence",
            "AI Model",
            "Assessor",
            "Assessed At",
            "Remediation Summary",
        ]
        _apply_headers(ws, headers)

        for row_idx, cr in enumerate(results, 2):
            ws.cell(row=row_idx, column=1, value=cr.id)
            ws.cell(row=row_idx, column=2, value=cr.framework)
            ws.cell(row=row_idx, column=3, value=cr.control_id)
            ws.cell(row=row_idx, column=4, value=cr.status)
            ws.cell(row=row_idx, column=5, value=cr.severity)
            ws.cell(row=row_idx, column=6, value=cr.assertion_name or "")
            ws.cell(
                row=row_idx,
                column=7,
                value=str(cr.assertion_passed) if cr.assertion_passed is not None else "",
            )
            ws.cell(row=row_idx, column=8, value=cr.ai_confidence)
            ws.cell(row=row_idx, column=9, value=cr.ai_model or "")
            ws.cell(row=row_idx, column=10, value=cr.assessor)
            ws.cell(row=row_idx, column=11, value=_iso_str(cr.assessed_at))
            ws.cell(row=row_idx, column=12, value=cr.remediation_summary or "")

        max_row = len(results) + 1
        if max_row > 1:
            _add_status_formatting(ws, 4, max_row)
            _add_severity_formatting(ws, 5, max_row)

        _auto_width(ws)
        ws.auto_filter.ref = f"A1:L{max_row}"

        out = Path(filepath)
        wb.save(str(out))
        log.info("Exported %d control results to %s", len(results), out)
        return out

    def export_poams(
        self,
        session: Session,
        filepath: str | Path,
        framework: str | None = None,
    ) -> Path:
        """Export POA&Ms with status coloring.

        Args:
            session: SQLAlchemy session.
            filepath: Output .xlsx path.
            framework: Optional framework filter.

        Returns:
            Path to the written file.
        """
        from warlock.db.models import POAM

        query = session.query(POAM)
        if framework:
            query = query.filter(POAM.framework == framework)
        query = query.order_by(POAM.created_at.desc())
        poams = query.limit(50000).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "POAMs"

        headers = [
            "ID",
            "Framework",
            "Control ID",
            "Weakness Description",
            "Severity",
            "Risk Level",
            "Status",
            "Scheduled Completion",
            "Actual Completion",
            "Cost Estimate",
            "Resource Allocation",
            "Escalation Level",
            "Created By",
            "Approved By",
            "Created At",
        ]
        _apply_headers(ws, headers)

        for row_idx, p in enumerate(poams, 2):
            ws.cell(row=row_idx, column=1, value=p.id)
            ws.cell(row=row_idx, column=2, value=p.framework)
            ws.cell(row=row_idx, column=3, value=p.control_id)
            ws.cell(row=row_idx, column=4, value=p.weakness_description)
            ws.cell(row=row_idx, column=5, value=p.severity)
            ws.cell(row=row_idx, column=6, value=p.risk_level or "")
            ws.cell(row=row_idx, column=7, value=p.status)
            ws.cell(row=row_idx, column=8, value=_iso_str(p.scheduled_completion))
            ws.cell(row=row_idx, column=9, value=_iso_str(p.actual_completion))
            cost_cell = ws.cell(row=row_idx, column=10, value=p.cost_estimate)
            if p.cost_estimate is not None:
                cost_cell.number_format = _CURRENCY_FORMAT
            ws.cell(row=row_idx, column=11, value=p.resource_allocation or "")
            ws.cell(row=row_idx, column=12, value=p.escalation_level or 0)
            ws.cell(row=row_idx, column=13, value=p.created_by or "")
            ws.cell(row=row_idx, column=14, value=p.approved_by or "")
            ws.cell(row=row_idx, column=15, value=_iso_str(p.created_at))

        max_row = len(poams) + 1
        if max_row > 1:
            _add_severity_formatting(ws, 5, max_row)
            # Status formatting for POA&M lifecycle
            status_col = get_column_letter(7)
            cell_range = f"{status_col}2:{status_col}{max_row}"
            ws.conditional_formatting.add(
                cell_range,
                CellIsRule(
                    operator="equal",
                    formula=['"closed"'],
                    fill=_GREEN_FILL,
                    font=_GREEN_FONT,
                ),
            )
            ws.conditional_formatting.add(
                cell_range,
                CellIsRule(
                    operator="equal",
                    formula=['"verified"'],
                    fill=_GREEN_FILL,
                    font=_GREEN_FONT,
                ),
            )
            ws.conditional_formatting.add(
                cell_range,
                CellIsRule(
                    operator="equal",
                    formula=['"draft"'],
                    fill=_YELLOW_FILL,
                    font=_YELLOW_FONT,
                ),
            )
            ws.conditional_formatting.add(
                cell_range,
                CellIsRule(
                    operator="equal",
                    formula=['"open"'],
                    fill=_RED_FILL,
                    font=_RED_FONT,
                ),
            )

        _auto_width(ws)
        ws.auto_filter.ref = f"A1:O{max_row}"

        out = Path(filepath)
        wb.save(str(out))
        log.info("Exported %d POAMs to %s", len(poams), out)
        return out

    def export_risk_register(
        self,
        session: Session,
        filepath: str | Path,
    ) -> Path:
        """Export risk analyses to an Excel workbook.

        Args:
            session: SQLAlchemy session.
            filepath: Output .xlsx path.

        Returns:
            Path to the written file.
        """
        from warlock.db.models import RiskAnalysis

        risks = (
            session.query(RiskAnalysis).order_by(RiskAnalysis.created_at.desc()).limit(50000).all()
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Risk Register"

        headers = [
            "ID",
            "Framework",
            "Scenario",
            "Mean ALE",
            "VaR 95%",
            "VaR 99%",
            "Control Effectiveness",
            "Iterations",
            "Risk Culture Score",
            "MTTR (days)",
            "Created At",
        ]
        _apply_headers(ws, headers)

        for row_idx, r in enumerate(risks, 2):
            ws.cell(row=row_idx, column=1, value=r.id)
            ws.cell(row=row_idx, column=2, value=r.framework)
            ws.cell(row=row_idx, column=3, value=r.scenario_name)
            ale_cell = ws.cell(row=row_idx, column=4, value=r.mean_ale)
            ale_cell.number_format = _CURRENCY_FORMAT
            var95_cell = ws.cell(row=row_idx, column=5, value=r.var_95)
            var95_cell.number_format = _CURRENCY_FORMAT
            var99_cell = ws.cell(row=row_idx, column=6, value=r.var_99)
            var99_cell.number_format = _CURRENCY_FORMAT
            ws.cell(row=row_idx, column=7, value=r.control_effectiveness)
            ws.cell(row=row_idx, column=8, value=r.iterations)
            ws.cell(row=row_idx, column=9, value=r.risk_culture_score)
            ws.cell(row=row_idx, column=10, value=r.mttr_days)
            ws.cell(row=row_idx, column=11, value=_iso_str(r.created_at))

        _auto_width(ws)
        max_row = len(risks) + 1
        ws.auto_filter.ref = f"A1:K{max_row}"

        out = Path(filepath)
        wb.save(str(out))
        log.info("Exported %d risk analyses to %s", len(risks), out)
        return out

    def export_generic(
        self,
        data: list[dict[str, Any]],
        filepath: str | Path,
        sheet_name: str = "Data",
    ) -> Path:
        """Export a generic list of dicts to an Excel workbook.

        Args:
            data: List of dictionaries to export.
            filepath: Output .xlsx path.
            sheet_name: Name for the worksheet.

        Returns:
            Path to the written file.
        """
        if not data:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            ws.cell(row=1, column=1, value="No data")
            out = Path(filepath)
            wb.save(str(out))
            return out

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        headers = list(data[0].keys())
        _apply_headers(ws, headers)

        for row_idx, row_data in enumerate(data, 2):
            for col_idx, key in enumerate(headers, 1):
                value = row_data.get(key)
                if isinstance(value, datetime):
                    value = _iso_str(value)
                elif isinstance(value, (list, dict)):
                    value = str(value)
                ws.cell(row=row_idx, column=col_idx, value=value)

        _auto_width(ws)
        max_row = len(data) + 1
        max_col = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A1:{max_col}{max_row}"

        out = Path(filepath)
        wb.save(str(out))
        log.info("Exported %d rows to %s (sheet: %s)", len(data), out, sheet_name)
        return out

    def export_multi_sheet(
        self,
        session: Session,
        filepath: str | Path,
        framework: str | None = None,
    ) -> Path:
        """Export a multi-sheet workbook with findings, control results, POAMs, and summary.

        Args:
            session: SQLAlchemy session.
            filepath: Output .xlsx path.
            framework: Optional framework filter applied to all sheets.

        Returns:
            Path to the written file.
        """
        from warlock.db.models import ControlResult, Finding, POAM, RiskAnalysis
        from sqlalchemy import func

        wb = Workbook()

        # --- Summary sheet ---
        ws_summary = wb.active
        ws_summary.title = "Summary"

        now = datetime.now(timezone.utc)
        summary_headers = ["Metric", "Value"]
        _apply_headers(ws_summary, summary_headers)

        q_findings = session.query(Finding)
        q_results = session.query(ControlResult)
        q_poams = session.query(POAM)
        q_risks = session.query(RiskAnalysis)

        if framework:
            finding_ids = (
                session.query(ControlResult.finding_id)
                .filter(ControlResult.framework == framework)
                .distinct()
            )
            q_findings = q_findings.filter(Finding.id.in_(finding_ids))
            q_results = q_results.filter(ControlResult.framework == framework)
            q_poams = q_poams.filter(POAM.framework == framework)
            q_risks = q_risks.filter(RiskAnalysis.framework == framework)

        finding_count = q_findings.count()
        result_count = q_results.count()
        poam_count = q_poams.count()
        risk_count = q_risks.count()

        compliant_count = q_results.filter(ControlResult.status == "compliant").count()
        non_compliant_count = q_results.filter(ControlResult.status == "non_compliant").count()
        pass_rate = f"{compliant_count / result_count * 100:.1f}%" if result_count else "N/A"

        fw_count = session.query(func.count(func.distinct(ControlResult.framework))).scalar()

        summary_rows = [
            ("Report Generated", now.strftime("%Y-%m-%d %H:%M UTC")),
            ("Framework Filter", framework or "All"),
            ("Total Findings", finding_count),
            ("Total Control Results", result_count),
            ("Compliant Results", compliant_count),
            ("Non-Compliant Results", non_compliant_count),
            ("Pass Rate", pass_rate),
            ("Active POAMs", poam_count),
            ("Risk Analyses", risk_count),
            ("Frameworks Assessed", fw_count),
        ]

        for row_idx, (metric, value) in enumerate(summary_rows, 2):
            ws_summary.cell(row=row_idx, column=1, value=metric)
            ws_summary.cell(row=row_idx, column=2, value=value)

        _auto_width(ws_summary)

        # --- Findings sheet ---
        ws_findings = wb.create_sheet("Findings")
        finding_headers = [
            "ID",
            "Title",
            "Observation Type",
            "Severity",
            "Source",
            "Provider",
            "Resource ID",
            "Observed At",
        ]
        _apply_headers(ws_findings, finding_headers)

        findings = q_findings.order_by(Finding.ingested_at.desc()).limit(50000).all()
        for row_idx, f in enumerate(findings, 2):
            ws_findings.cell(row=row_idx, column=1, value=f.id)
            ws_findings.cell(row=row_idx, column=2, value=f.title)
            ws_findings.cell(row=row_idx, column=3, value=f.observation_type)
            ws_findings.cell(row=row_idx, column=4, value=f.severity)
            ws_findings.cell(row=row_idx, column=5, value=f.source)
            ws_findings.cell(row=row_idx, column=6, value=f.provider)
            ws_findings.cell(row=row_idx, column=7, value=f.resource_id or "")
            ws_findings.cell(row=row_idx, column=8, value=_iso_str(f.observed_at))

        f_max = len(findings) + 1
        if f_max > 1:
            _add_severity_formatting(ws_findings, 4, f_max)
        _auto_width(ws_findings)
        ws_findings.auto_filter.ref = f"A1:H{f_max}"

        # --- Control Results sheet ---
        ws_results = wb.create_sheet("Control Results")
        cr_headers = [
            "Framework",
            "Control ID",
            "Status",
            "Severity",
            "Assertion",
            "Assessor",
            "Assessed At",
        ]
        _apply_headers(ws_results, cr_headers)

        results = (
            q_results.order_by(ControlResult.framework, ControlResult.control_id)
            .limit(100000)
            .all()
        )
        for row_idx, cr in enumerate(results, 2):
            ws_results.cell(row=row_idx, column=1, value=cr.framework)
            ws_results.cell(row=row_idx, column=2, value=cr.control_id)
            ws_results.cell(row=row_idx, column=3, value=cr.status)
            ws_results.cell(row=row_idx, column=4, value=cr.severity)
            ws_results.cell(row=row_idx, column=5, value=cr.assertion_name or "")
            ws_results.cell(row=row_idx, column=6, value=cr.assessor)
            ws_results.cell(row=row_idx, column=7, value=_iso_str(cr.assessed_at))

        cr_max = len(results) + 1
        if cr_max > 1:
            _add_status_formatting(ws_results, 3, cr_max)
            _add_severity_formatting(ws_results, 4, cr_max)
        _auto_width(ws_results)
        ws_results.auto_filter.ref = f"A1:G{cr_max}"

        # --- POAMs sheet ---
        ws_poams = wb.create_sheet("POAMs")
        poam_headers = [
            "Framework",
            "Control ID",
            "Weakness",
            "Severity",
            "Status",
            "Scheduled Completion",
            "Cost Estimate",
            "Created By",
        ]
        _apply_headers(ws_poams, poam_headers)

        poams = q_poams.order_by(POAM.created_at.desc()).limit(50000).all()
        for row_idx, p in enumerate(poams, 2):
            ws_poams.cell(row=row_idx, column=1, value=p.framework)
            ws_poams.cell(row=row_idx, column=2, value=p.control_id)
            ws_poams.cell(row=row_idx, column=3, value=p.weakness_description)
            ws_poams.cell(row=row_idx, column=4, value=p.severity)
            ws_poams.cell(row=row_idx, column=5, value=p.status)
            ws_poams.cell(row=row_idx, column=6, value=_iso_str(p.scheduled_completion))
            cost_cell = ws_poams.cell(row=row_idx, column=7, value=p.cost_estimate)
            if p.cost_estimate is not None:
                cost_cell.number_format = _CURRENCY_FORMAT
            ws_poams.cell(row=row_idx, column=8, value=p.created_by or "")

        p_max = len(poams) + 1
        if p_max > 1:
            _add_severity_formatting(ws_poams, 4, p_max)
        _auto_width(ws_poams)
        ws_poams.auto_filter.ref = f"A1:H{p_max}"

        out = Path(filepath)
        wb.save(str(out))
        log.info(
            "Exported multi-sheet workbook to %s: %d findings, %d results, %d POAMs",
            out,
            len(findings),
            len(results),
            len(poams),
        )
        return out
